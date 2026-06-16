"""MIRA exports — Excel/CSV download of Atlas CMMS data.

Endpoints (mounted on the mira-mcp REST app):
    GET /api/v1/exports/assets.xlsx          — all assets for authenticated tenant
    GET /api/v1/exports/assets.xlsx?format=csv
    GET /api/v1/exports/work-orders.xlsx     — all work orders for authenticated tenant
    GET /api/v1/exports/work-orders.xlsx?format=csv

Auth: PLG JWT (HS256, ``PLG_JWT_SECRET`` env var) in ``Authorization: Bearer <token>``
      or ``?token=`` query param or ``mira_session`` cookie.
      JWT ``sub`` claim is the tenant_id used for all Atlas scoping.

Streaming: rows are fetched in batches of 1000 to avoid loading 50k+ WOs into memory.

Column names match csv-import.ts (mira-web/src/lib/csv-import.ts) for round-trip parity.

Licenses used:
    openpyxl 3.1.5  — MIT
    PyJWT   >=2.8   — MIT
"""

from __future__ import annotations

import csv
import io
import logging
import os
from typing import AsyncIterator

import jwt as _pyjwt
import openpyxl
from cmms.atlas import AtlasCMMS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from starlette.requests import Request
from starlette.responses import Response
from tenant_resolver import resolve_atlas_creds

logger = logging.getLogger("mira-exports")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BATCH_SIZE = 1000  # rows fetched per Atlas page request
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# ---------------------------------------------------------------------------
# JWT verification
# ---------------------------------------------------------------------------


def _get_jwt_secret() -> str:
    secret = os.environ.get("PLG_JWT_SECRET", "")
    if not secret:
        logger.warning("PLG_JWT_SECRET not set — export auth will always reject")
    return secret


def _extract_jwt(request: Request) -> str | None:
    """Extract raw JWT from Authorization header, ?token= param, or cookie."""
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[len("Bearer ") :]
    token_param = request.query_params.get("token")
    if token_param:
        return token_param
    cookie = request.cookies.get("mira_session")
    return cookie or None


def _verify_jwt(token: str) -> dict | None:
    """Return decoded payload dict or None on any verification failure."""
    secret = _get_jwt_secret()
    if not secret:
        return None
    try:
        return _pyjwt.decode(token, secret, algorithms=["HS256"])
    except _pyjwt.ExpiredSignatureError:
        logger.info("JWT expired")
    except _pyjwt.InvalidTokenError as exc:
        logger.info("JWT invalid: %s", exc)
    return None


def _require_tenant(request: Request) -> tuple[str, None] | tuple[None, Response]:
    """
    Extract and verify JWT. Return ``(tenant_id, None)`` on success or
    ``(None, error_response)`` on failure.
    """
    raw = _extract_jwt(request)
    if not raw:
        return None, Response(
            '{"error":"Unauthorized"}', status_code=401, media_type="application/json"
        )

    payload = _verify_jwt(raw)
    if not payload:
        return None, Response(
            '{"error":"Invalid or expired token"}', status_code=401, media_type="application/json"
        )

    tenant_id = payload.get("sub", "")
    if not tenant_id:
        return None, Response(
            '{"error":"Token missing sub claim"}', status_code=401, media_type="application/json"
        )

    return tenant_id, None


# ---------------------------------------------------------------------------
# Atlas helpers — streamed batch fetching
# ---------------------------------------------------------------------------


async def _atlas_for_tenant(tenant_id: str) -> AtlasCMMS | None:
    """Resolve tenant Atlas credentials and return a ready client, or None."""
    creds = await resolve_atlas_creds(tenant_id)
    if not creds:
        logger.warning("No Atlas creds for tenant=%s", tenant_id)
        return None
    email, password, api_url = creds
    return AtlasCMMS.for_tenant(email, password, api_url)


async def _stream_work_orders(atlas: AtlasCMMS) -> AsyncIterator[dict]:
    """Yield work order dicts in batches of BATCH_SIZE, stopping when Atlas returns empty."""
    page = 0
    while True:
        try:
            token = await atlas._get_token()
            if not token:
                break
            async with __import__("httpx").AsyncClient(timeout=60) as client:
                resp = await client.post(
                    f"{atlas.api_url}/work-orders/search",
                    headers={
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/json",
                    },
                    json={"pageSize": BATCH_SIZE, "pageNum": page},
                )
                resp.raise_for_status()
                data = resp.json()
        except Exception as exc:
            logger.error("Atlas work-orders page=%d failed: %s", page, exc)
            break

        rows: list[dict] = data if isinstance(data, list) else data.get("content", [])
        if not rows:
            break
        for row in rows:
            yield row
        if len(rows) < BATCH_SIZE:
            break
        page += 1


async def _stream_assets(atlas: AtlasCMMS) -> AsyncIterator[dict]:
    """Yield asset dicts in batches of BATCH_SIZE."""
    page = 0
    while True:
        try:
            token = await atlas._get_token()
            if not token:
                break
            async with __import__("httpx").AsyncClient(timeout=60) as client:
                resp = await client.post(
                    f"{atlas.api_url}/assets/search",
                    headers={
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/json",
                    },
                    json={"pageSize": BATCH_SIZE, "pageNum": page},
                )
                resp.raise_for_status()
                data = resp.json()
        except Exception as exc:
            logger.error("Atlas assets page=%d failed: %s", page, exc)
            break

        rows: list[dict] = data if isinstance(data, list) else data.get("content", [])
        if not rows:
            break
        for row in rows:
            yield row
        if len(rows) < BATCH_SIZE:
            break
        page += 1


# ---------------------------------------------------------------------------
# Row formatters  (column names match csv-import.ts for round-trip parity)
# ---------------------------------------------------------------------------

# Work-order columns mirror csv-import.ts CSVRow interface:
#   date, title, description, priority, asset, category
# Additional Atlas fields appended after the importable columns.
WORK_ORDER_HEADERS = [
    "id",
    "date",
    "title",
    "description",
    "priority",
    "asset",
    "category",
    "status",
]

# Asset columns match Atlas asset structure + what csv-import.ts uses for asset field
ASSET_HEADERS = [
    "id",
    "name",
    "description",
    "manufacturer",
    "model",
    "serialNumber",
    "status",
    "location",
    "createdAt",
]


def format_work_order_row(wo: dict) -> list[str]:
    """Extract work-order fields into a flat list matching WORK_ORDER_HEADERS."""
    asset_ref = wo.get("asset") or {}
    asset_name = asset_ref.get("name", "") if isinstance(asset_ref, dict) else str(asset_ref)
    return [
        str(wo.get("id", "")),
        str(wo.get("createdAt", wo.get("created_at", wo.get("dueDate", "")))),
        str(wo.get("title", "")),
        str(wo.get("description", "")),
        str(wo.get("priority", "MEDIUM")),
        asset_name,
        str(wo.get("category", "")),
        str(wo.get("status", "")),
    ]


def format_asset_row(asset: dict) -> list[str]:
    """Extract asset fields into a flat list matching ASSET_HEADERS."""
    return [
        str(asset.get("id", "")),
        str(asset.get("name", "")),
        str(asset.get("description", "")),
        str(asset.get("manufacturer", "")),
        str(asset.get("model", "")),
        str(asset.get("serialNumber", asset.get("serial_number", ""))),
        str(asset.get("status", "")),
        str(asset.get("location", asset.get("locationPath", ""))),
        str(asset.get("createdAt", asset.get("created_at", ""))),
    ]


# ---------------------------------------------------------------------------
# Workbook builder helpers
# ---------------------------------------------------------------------------


def _style_header_row(ws, headers: list[str]) -> None:
    """Write and style the header row in the active worksheet."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = max(15, len(str(cell.value)) + 4)


async def _build_xlsx(
    headers: list[str],
    row_iter: AsyncIterator[dict],
    formatter,
    sheet_name: str,
) -> bytes:
    """Build an in-memory .xlsx workbook by consuming row_iter, return bytes."""
    wb: Workbook = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    _style_header_row(ws, headers)

    async for record in row_iter:
        ws.append(formatter(record))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _build_csv(
    headers: list[str],
    row_iter: AsyncIterator[dict],
    formatter,
) -> bytes:
    """Build an in-memory CSV by consuming row_iter, return UTF-8 bytes with BOM.

    The BOM (utf-8-sig) ensures Excel opens the file without a "data loss" warning
    and renders accented characters correctly.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow(headers)

    async for record in row_iter:
        writer.writerow(formatter(record))

    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Route handlers
# ---------------------------------------------------------------------------


async def export_work_orders(request: Request) -> Response:
    """GET /api/v1/exports/work-orders.xlsx  (or ?format=csv)

    Returns all work orders for the authenticated tenant's Atlas account.
    Streams in batches of 1000 rows from Atlas — safe for 50k+ WOs.
    """
    tenant_id, err = _require_tenant(request)
    if err is not None:
        return err

    atlas = await _atlas_for_tenant(tenant_id)
    if atlas is None:
        return Response(
            '{"error":"Atlas CMMS not configured for this tenant"}',
            status_code=503,
            media_type="application/json",
        )

    fmt = request.query_params.get("format", "xlsx").lower()
    row_iter = _stream_work_orders(atlas)

    if fmt == "csv":
        body = await _build_csv(WORK_ORDER_HEADERS, row_iter, format_work_order_row)
        return Response(
            content=body,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="work-orders.csv"',
                "Cache-Control": "no-store",
            },
        )

    # Default: xlsx
    body = await _build_xlsx(
        WORK_ORDER_HEADERS,
        row_iter,
        format_work_order_row,
        "Work Orders",
    )
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="work-orders.xlsx"',
            "Cache-Control": "no-store",
        },
    )


async def export_assets(request: Request) -> Response:
    """GET /api/v1/exports/assets.xlsx  (or ?format=csv)

    Returns all assets for the authenticated tenant's Atlas account.
    """
    tenant_id, err = _require_tenant(request)
    if err is not None:
        return err

    atlas = await _atlas_for_tenant(tenant_id)
    if atlas is None:
        return Response(
            '{"error":"Atlas CMMS not configured for this tenant"}',
            status_code=503,
            media_type="application/json",
        )

    fmt = request.query_params.get("format", "xlsx").lower()
    row_iter = _stream_assets(atlas)

    if fmt == "csv":
        body = await _build_csv(ASSET_HEADERS, row_iter, format_asset_row)
        return Response(
            content=body,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="assets.csv"',
                "Cache-Control": "no-store",
            },
        )

    body = await _build_xlsx(
        ASSET_HEADERS,
        row_iter,
        format_asset_row,
        "Assets",
    )
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="assets.xlsx"',
            "Cache-Control": "no-store",
        },
    )
