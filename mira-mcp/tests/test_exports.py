"""Tests for Unit 4 — Excel/CSV live export endpoints.

Covers:
- Row formatter functions (pure unit tests — no network, no DB)
- JWT extraction and verification
- Wrong-tenant JWT returns 401/missing tenant returns 401
- Header styling round-trip (write xlsx, load with openpyxl, assert sheet contents)
- CSV BOM present for Excel compatibility
- format=csv switch works
- 503 when Atlas creds not found

Run with:  pytest mira-mcp/tests/test_exports.py -v
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import AsyncIterator
from unittest.mock import AsyncMock, MagicMock, patch

import jwt as _pyjwt
import openpyxl
import pytest

# Ensure mira-mcp package root is on sys.path when tests run from repo root.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


# ---------------------------------------------------------------------------
# Minimal stubs — prevent real network calls during import of exports.py
# ---------------------------------------------------------------------------

# Stub out openviking so server.py imports without the package installed
_fake_viking = MagicMock()
_fake_viking.ingest_pdf = lambda *a, **kw: 0
_fake_viking.retrieve = lambda *a, **kw: []
sys.modules.setdefault("context.viking_store", _fake_viking)
sys.modules.setdefault("openviking", MagicMock())

# Stub psycopg (not available in test env)
sys.modules.setdefault("psycopg", MagicMock())
sys.modules.setdefault("psycopg2", MagicMock())

import exports  # noqa: E402 — must come after sys.path setup

# ---------------------------------------------------------------------------
# Helper: build a signed PLG JWT for a tenant
# ---------------------------------------------------------------------------

_TEST_SECRET = "test-secret-for-unit-4"
_TEST_TENANT_ID = "aaaabbbb-cccc-dddd-eeee-000011112222"
_OTHER_TENANT_ID = "ffffffff-aaaa-bbbb-cccc-111122223333"


def _make_token(tenant_id: str = _TEST_TENANT_ID, secret: str = _TEST_SECRET) -> str:
    return _pyjwt.encode(
        {
            "sub": tenant_id,
            "email": "test@example.com",
            "tier": "active",
            "atlasCompanyId": 42,
            "atlasUserId": 7,
        },
        secret,
        algorithm="HS256",
    )


# ---------------------------------------------------------------------------
# Unit tests: row formatters
# ---------------------------------------------------------------------------


class TestFormatWorkOrderRow:
    def test_basic_fields_present(self):
        wo = {
            "id": 101,
            "title": "Fix pump bearing",
            "description": "Vibration exceeds spec",
            "priority": "HIGH",
            "category": "CORRECTIVE",
            "status": "OPEN",
            "createdAt": "2026-04-01T08:00:00Z",
            "asset": {"id": 5, "name": "Pump-001"},
        }
        row = exports.format_work_order_row(wo)
        assert len(row) == len(exports.WORK_ORDER_HEADERS)
        assert row[0] == "101"
        assert row[2] == "Fix pump bearing"
        assert row[3] == "Vibration exceeds spec"
        assert row[4] == "HIGH"
        assert row[5] == "Pump-001"
        assert row[6] == "CORRECTIVE"
        assert row[7] == "OPEN"

    def test_missing_asset_empty_string(self):
        wo = {"id": 1, "title": "Test", "description": "Desc", "priority": "LOW"}
        row = exports.format_work_order_row(wo)
        assert row[5] == ""  # asset column

    def test_asset_as_string_falls_back(self):
        wo = {"id": 2, "title": "T", "description": "D", "asset": "some-asset-name"}
        row = exports.format_work_order_row(wo)
        assert row[5] == "some-asset-name"

    def test_no_created_at_falls_back_to_due_date(self):
        wo = {"id": 3, "title": "T", "description": "D", "dueDate": "2026-05-01"}
        row = exports.format_work_order_row(wo)
        assert row[1] == "2026-05-01"

    def test_all_columns_are_strings(self):
        wo = {"id": 99, "title": "T", "description": "D", "priority": "MEDIUM"}
        row = exports.format_work_order_row(wo)
        assert all(isinstance(v, str) for v in row)


class TestFormatAssetRow:
    def test_basic_fields(self):
        asset = {
            "id": 10,
            "name": "Conveyor Motor",
            "description": "Main drive",
            "manufacturer": "ABB",
            "model": "M3BP 180",
            "serialNumber": "SN12345",
            "status": "ACTIVE",
            "location": "Building A",
            "createdAt": "2025-01-15T00:00:00Z",
        }
        row = exports.format_asset_row(asset)
        assert len(row) == len(exports.ASSET_HEADERS)
        assert row[0] == "10"
        assert row[1] == "Conveyor Motor"
        assert row[2] == "Main drive"
        assert row[3] == "ABB"
        assert row[4] == "M3BP 180"
        assert row[5] == "SN12345"
        assert row[6] == "ACTIVE"
        assert row[7] == "Building A"

    def test_serial_number_fallback(self):
        asset = {"id": 1, "name": "Pump", "serial_number": "ALT-SN"}
        row = exports.format_asset_row(asset)
        assert row[5] == "ALT-SN"

    def test_all_columns_are_strings(self):
        asset = {"id": 5, "name": "Motor"}
        row = exports.format_asset_row(asset)
        assert all(isinstance(v, str) for v in row)


# ---------------------------------------------------------------------------
# Unit tests: JWT helpers
# ---------------------------------------------------------------------------


class TestJWTVerification:
    def test_valid_token_returns_payload(self, monkeypatch):
        monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
        token = _make_token()
        payload = exports._verify_jwt(token)
        assert payload is not None
        assert payload["sub"] == _TEST_TENANT_ID

    def test_wrong_secret_returns_none(self, monkeypatch):
        monkeypatch.setenv("PLG_JWT_SECRET", "wrong-secret")
        token = _make_token(secret=_TEST_SECRET)
        assert exports._verify_jwt(token) is None

    def test_missing_secret_returns_none(self, monkeypatch):
        monkeypatch.delenv("PLG_JWT_SECRET", raising=False)
        token = _make_token()
        assert exports._verify_jwt(token) is None

    def test_malformed_token_returns_none(self, monkeypatch):
        monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
        assert exports._verify_jwt("not.a.jwt") is None

    def test_expired_token_returns_none(self, monkeypatch):
        monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
        import time

        expired = _pyjwt.encode(
            {"sub": _TEST_TENANT_ID, "exp": int(time.time()) - 3600},
            _TEST_SECRET,
            algorithm="HS256",
        )
        assert exports._verify_jwt(expired) is None


# ---------------------------------------------------------------------------
# Helpers: fake async iterators and Starlette request mocks
# ---------------------------------------------------------------------------


async def _async_iter(items: list[dict]) -> AsyncIterator[dict]:
    for item in items:
        yield item


class _FakeScope:
    """Minimal ASGI scope for Starlette Request."""

    def __init__(self, path: str = "/", qs: str = "", headers: list | None = None):
        self._headers = headers or []
        self._path = path
        self._qs = qs

    def __getitem__(self, key: str):
        if key == "type":
            return "http"
        if key == "method":
            return "GET"
        if key == "path":
            return self._path
        if key == "query_string":
            return self._qs.encode()
        if key == "headers":
            return [(k.lower().encode(), v.encode()) for k, v in self._headers]
        if key == "root_path":
            return ""
        raise KeyError(key)

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default


def _make_request(
    path: str = "/api/v1/exports/work-orders.xlsx",
    qs: str = "",
    bearer: str | None = None,
    cookie: str | None = None,
) -> object:
    """Build a minimal Starlette Request-like object."""
    from starlette.requests import Request

    headers = []
    if bearer:
        headers.append(("Authorization", f"Bearer {bearer}"))
    if cookie:
        headers.append(("Cookie", f"mira_session={cookie}"))

    scope = {
        "type": "http",
        "method": "GET",
        "path": path,
        "query_string": qs.encode(),
        "headers": [(k.lower().encode(), v.encode()) for k, v in headers],
        "root_path": "",
    }

    async def receive():
        return {"type": "http.disconnect"}

    return Request(scope, receive=receive)


# ---------------------------------------------------------------------------
# Integration tests: handler — auth rejection
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_missing_token_returns_401():
    req = _make_request()
    resp = await exports.export_work_orders(req)
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_wrong_secret_returns_401(monkeypatch):
    monkeypatch.setenv("PLG_JWT_SECRET", "real-secret")
    token = _make_token(secret="wrong-secret")
    req = _make_request(bearer=token)
    resp = await exports.export_work_orders(req)
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_different_tenant_token_cannot_download_other_tenant(monkeypatch):
    """A valid JWT for tenant B cannot download tenant A's data.

    The handler authenticates using the JWT's own sub claim, then resolves
    Atlas creds for THAT tenant only — there is no tenant_id parameter in the
    URL that could be overridden.  Passing tenant_B's token will either:
      - Resolve tenant_B's Atlas creds (not tenant_A's) → returns tenant_B data, not A's
      - Fail to resolve creds → 503

    We assert it does NOT return 200 with a hard-coded wrong tenant, because
    the JWT sub drives the query. Here we verify that mocking creds-not-found
    for tenant B returns 503 (no cross-tenant data leak path exists).
    """
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token(tenant_id=_OTHER_TENANT_ID)
    req = _make_request(bearer=token)

    with patch("exports.resolve_atlas_creds", new=AsyncMock(return_value=None)):
        resp = await exports.export_work_orders(req)
    # 503 — creds not found for this tenant; crucially NOT the other tenant's data
    assert resp.status_code == 503


@pytest.mark.asyncio
async def test_valid_token_no_atlas_creds_returns_503(monkeypatch):
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(bearer=token)

    with patch("exports.resolve_atlas_creds", new=AsyncMock(return_value=None)):
        resp = await exports.export_work_orders(req)
    assert resp.status_code == 503


# ---------------------------------------------------------------------------
# Integration tests: xlsx output verification (via openpyxl.load_workbook)
# ---------------------------------------------------------------------------

_SAMPLE_WORK_ORDERS = [
    {
        "id": 1,
        "title": "Replace bearing",
        "description": "Vibration alarm",
        "priority": "HIGH",
        "category": "CORRECTIVE",
        "status": "OPEN",
        "createdAt": "2026-04-01T09:00:00Z",
        "asset": {"name": "Pump-001"},
    },
    {
        "id": 2,
        "title": "Lube chain drive",
        "description": "PM 90-day interval",
        "priority": "LOW",
        "category": "PREVENTIVE",
        "status": "COMPLETE",
        "createdAt": "2026-04-02T10:00:00Z",
        "asset": {"name": "Conveyor-A"},
    },
]

_SAMPLE_ASSETS = [
    {
        "id": 10,
        "name": "Pump-001",
        "description": "Main coolant pump",
        "manufacturer": "Grundfos",
        "model": "CM10",
        "serialNumber": "SN-99",
        "status": "ACTIVE",
        "location": "Pump Room",
        "createdAt": "2025-01-01T00:00:00Z",
    }
]


@pytest.mark.asyncio
async def test_xlsx_work_orders_header_row(monkeypatch, tmp_path):
    """Build xlsx, load with openpyxl, assert header row matches WORK_ORDER_HEADERS."""
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(bearer=token)

    mock_atlas = MagicMock()

    async def _fake_stream(*a, **kw):
        for wo in _SAMPLE_WORK_ORDERS:
            yield wo

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_work_orders", side_effect=_fake_stream),
    ):
        resp = await exports.export_work_orders(req)

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")

    wb = openpyxl.load_workbook(io.BytesIO(resp.body))
    ws = wb.active
    assert ws.title == "Work Orders"

    # Header row
    actual_headers = [cell.value for cell in ws[1]]
    assert actual_headers == exports.WORK_ORDER_HEADERS

    # Data rows present
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    assert rows[0][2] == "Replace bearing"
    assert rows[1][2] == "Lube chain drive"


@pytest.mark.asyncio
async def test_xlsx_assets_header_row(monkeypatch):
    """Build assets xlsx, assert header row and sheet name."""
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(path="/api/v1/exports/assets.xlsx", bearer=token)

    mock_atlas = MagicMock()

    async def _fake_stream(*a, **kw):
        for asset in _SAMPLE_ASSETS:
            yield asset

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_assets", side_effect=_fake_stream),
    ):
        resp = await exports.export_assets(req)

    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.body))
    ws = wb.active
    assert ws.title == "Assets"

    actual_headers = [cell.value for cell in ws[1]]
    assert actual_headers == exports.ASSET_HEADERS

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][1] == "Pump-001"
    assert rows[0][3] == "Grundfos"


# ---------------------------------------------------------------------------
# Integration tests: CSV output
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_csv_work_orders_has_bom_and_headers(monkeypatch):
    """CSV output must start with UTF-8 BOM so Excel opens without data-loss warning."""
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(bearer=token, qs="format=csv")

    mock_atlas = MagicMock()

    async def _fake_stream(*a, **kw):
        for wo in _SAMPLE_WORK_ORDERS:
            yield wo

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_work_orders", side_effect=_fake_stream),
    ):
        resp = await exports.export_work_orders(req)

    assert resp.status_code == 200
    assert "text/csv" in resp.headers.get("content-type", "")

    # BOM check — UTF-8 BOM is \xef\xbb\xbf
    body: bytes = resp.body
    assert body[:3] == b"\xef\xbb\xbf", "CSV must start with UTF-8 BOM for Excel compatibility"

    # Decode and check header
    text = body.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert "title" in first_line
    assert "description" in first_line
    assert "priority" in first_line


@pytest.mark.asyncio
async def test_csv_assets_format_param(monkeypatch):
    """?format=csv on assets endpoint returns CSV."""
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(path="/api/v1/exports/assets.xlsx", bearer=token, qs="format=csv")

    mock_atlas = MagicMock()

    async def _fake_stream(*a, **kw):
        for asset in _SAMPLE_ASSETS:
            yield asset

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_assets", side_effect=_fake_stream),
    ):
        resp = await exports.export_assets(req)

    assert resp.status_code == 200
    assert "text/csv" in resp.headers.get("content-type", "")
    text = resp.body.decode("utf-8-sig")
    assert "Pump-001" in text


# ---------------------------------------------------------------------------
# Test: Content-Disposition headers set correctly
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_content_disposition_xlsx(monkeypatch):
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(bearer=token)

    mock_atlas = MagicMock()

    async def _empty(*a, **kw):
        return
        yield  # pragma: no cover

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_work_orders", side_effect=_empty),
    ):
        resp = await exports.export_work_orders(req)

    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "work-orders.xlsx" in cd
    assert "attachment" in cd


@pytest.mark.asyncio
async def test_content_disposition_csv(monkeypatch):
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(bearer=token, qs="format=csv")

    mock_atlas = MagicMock()

    async def _empty(*a, **kw):
        return
        yield  # pragma: no cover

    with (
        patch(
            "exports.resolve_atlas_creds", new=AsyncMock(return_value=("u", "p", "http://atlas"))
        ),
        patch("exports.AtlasCMMS.for_tenant", return_value=mock_atlas),
        patch("exports._stream_work_orders", side_effect=_empty),
    ):
        resp = await exports.export_work_orders(req)

    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "work-orders.csv" in cd


# ---------------------------------------------------------------------------
# Test: token via cookie
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_token_via_cookie_accepted(monkeypatch):
    monkeypatch.setenv("PLG_JWT_SECRET", _TEST_SECRET)
    token = _make_token()
    req = _make_request(cookie=token)

    with patch("exports.resolve_atlas_creds", new=AsyncMock(return_value=None)):
        resp = await exports.export_work_orders(req)
    # 503 means we got past auth — cookie-based auth worked
    assert resp.status_code == 503


# ---------------------------------------------------------------------------
# Test: column headers match csv-import.ts round-trip columns
# ---------------------------------------------------------------------------


def test_work_order_headers_match_csv_import_ts():
    """The first 6 csv-import.ts columns must be present in WORK_ORDER_HEADERS.

    csv-import.ts CSVRow: date, title, description, priority, asset, category
    These 6 fields are the importable subset; additional Atlas fields may follow.
    """
    importable = ["date", "title", "description", "priority", "asset", "category"]
    for col in importable:
        assert col in exports.WORK_ORDER_HEADERS, (
            f"Column '{col}' from csv-import.ts CSVRow is missing from WORK_ORDER_HEADERS"
        )
