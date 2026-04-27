"""Build MIRA tech-debt register xlsx (one-off, run once)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUT = "/sessions/compassionate-bold-lovelace/mnt/MIRA/docs/tech-debt/2026-04-25-debt-register.xlsx"

EFFORT_NUM = {"S": 1, "M": 2, "L": 3, "XL": 5}

# 28 findings: (id, title, module, file_anchor, category, severity, effort, impact, risk, phase, why, fix)
FINDINGS = [
    ("DBT-001", "Supervisor god object", "mira-bots", "shared/engine.py (2531 LOC)", "Architecture", "Critical", "L", 5, 4, 3,
     "engine.py mixes FSM, 5 workers, guardrails, fallbacks, work-order formatting, photo persistence in 2531 lines. Adding any feature touches this file. Merge conflicts compound across the parity sprint.",
     "Extract FSMController, WorkerCoordinator, ResponseFormatter. Eval-gate each step. Schedule before file crosses 3000 lines."),
    ("DBT-002", "Three-layer LLM cascade with scattered failover", "mira-bots, mira-pipeline", "router.py:289; rag_worker.py:186; pipeline/main.py", "Architecture", "Critical", "M", 5, 5, 2,
     "Router.complete() owns one cascade; RAG worker has its own 3-stage pipeline; pipeline wraps engine which calls RAG which calls router. A failover bug fixed in one layer doesn't reach the others.",
     "Router owns cascade. RAGWorker calls router only. Pipeline calls engine only. Centralize telemetry at router level."),
    ("DBT-003", "Pipeline imports mira-bots via sys.path hack", "mira-pipeline", "mira-pipeline/main.py:32, 1107", "Architecture", "High", "M", 4, 4, 2,
     "sys.path.insert(0, ...) in two places. No declared dependency. Breaks packaging, deployment isolation, and type checking.",
     "Package mira-bots/shared as publishable library. Declare in mira-pipeline/requirements.txt. Remove sys.path manipulation."),
    ("DBT-004", "15+ scattered timeout literals", "mira-bots, mira-pipeline", "router.py 4x; engine.py 2x; nemotron.py; agents/*.py 5x; workers/*.py 3x", "Code", "High", "S", 4, 3, 1,
     "Timeouts (15/30/60/120s) hardcoded in 15+ httpx.AsyncClient and asyncio.wait_for calls. Provider-degradation incident requires hunting and editing many files.",
     "shared/config.py with TimeoutPolicy dataclass loaded at startup. All callers reference policy."),
    ("DBT-005", "_PROVIDER_HOURLY_LIMITS hardcoded in class dict", "mira-bots", "inference/router.py:218", "Architecture", "High", "S", 4, 3, 1,
     "Provider call budgets (groq:1800, cerebras:1800, gemini:900, claude:5000) baked into a class dict. Rate-limit incidents require code deploy.",
     "Load INFERENCE_PROVIDER_LIMITS from Doppler env (JSON or comma-separated). Default to current values."),
    ("DBT-006", "Provider cascade order hardcoded", "mira-bots", "inference/router.py:_build_providers", "Code", "Medium", "S", 3, 3, 1,
     "Provider order Groq -> Cerebras -> Gemini -> Claude is hardcoded. Comments explain historical reordering after Gemini 503s. Future reordering requires another deploy.",
     "INFERENCE_PROVIDER_ORDER env var (comma-separated). Parse in _build_providers. Default to current order."),
    ("DBT-007", "PII sanitizer not portable to mira-hub", "mira-bots, mira-hub", "router.py:42-50 only", "Architecture", "High", "M", 4, 5, 2,
     "Production IPv4/MAC/serial regex lives in mira-bots router.py. mira-hub asset-chat route has a 3-rule stub. PII can leak into chat responses on the new product surface.",
     "Port regex to mira-hub/src/lib/pii-sanitizer.ts. Wrap as sanitizeContext(text). Call before LLM dispatch in /api/v1/assets/[id]/chat."),
    ("DBT-008", "libsodium wrapping duplicated", "mira-hub", "lib/llm-keys.ts + lib/webhooks/secret-storage.ts", "Code", "Medium", "M", 3, 4, 2,
     "Both files independently implement getSodium loader, getKEK validation, nonce concatenation, stub mode. secret-storage.ts already has a TODO(#576) referencing this duplication.",
     "Extract to mira-hub/src/lib/crypto/secretbox.ts: encryptSecretbox/decryptSecretbox. Both callers update."),
    ("DBT-009", "KEK rotation story missing from #574", "mira-hub", "lib/llm-keys.ts (no key_id column)", "Architecture", "Critical", "M", 5, 5, 2,
     "If LLM_KEK rotates, every encrypted llm_keys row becomes unreadable. No key_id, no re-encryption migration path. Doing this after launch is a customer-impacting outage.",
     "Add key_id column referencing KEK version. Re-encryption migration path. Validate hex KEK at module load. Land before storing real customer keys."),
    ("DBT-010", "Tenant audit log allows UPDATE/DELETE under FOR ALL policy", "mira-hub", "db/migrations/2026-04-24-008-tenants-rls.sql", "Architecture", "Critical", "S", 5, 5, 1,
     "FOR ALL policy on tenant_audit_log permits app-role UPDATE/DELETE. SOC 2 CC4.1 / CC7.2 require audit-log immutability. Blocks SOC 2 Type 1 evidence collection.",
     "Split into FOR INSERT WITH CHECK (tenant_id_match) and FOR SELECT USING (tenant_id_match). No update/delete grant for app role."),
    ("DBT-011", "guardrails.py 986 LOC mixing 5 concerns", "mira-bots", "shared/guardrails.py (986 LOC)", "Architecture", "Medium", "L", 3, 3, 3,
     "Single module owns SAFETY_KEYWORDS, classify_intent, check_output, resolve_option_selection, strip_mentions, vendor_name_from_text, vendor_support_url. Can't test vendor logic without safety logic.",
     "Split into intent.py, safety.py, vendors.py, text_cleaning.py, option_resolution.py. Re-export from guardrails/__init__.py for compat."),
    ("DBT-012", "ingest main.py 1621 LOC", "mira-core", "mira-core/mira-ingest/main.py (1621 LOC)", "Architecture", "Medium", "L", 3, 3, 3,
     "One 1.6k-line file handles photo ingest, visual search, document KB, Firecrawl, Apify, crawl verification, Reddit benchmark, Telegram notifications. Single point of change.",
     "Refactor into routes/, pipelines/, services/. Example: routes/photo.py, services/scrape.py, services/verify.py. main.py keeps FastAPI setup only."),
    ("DBT-013", "pipeline main.py 1331 LOC (was ~500 on April 14)", "mira-pipeline", "mira-pipeline/main.py (1331 LOC)", "Code", "Medium", "M", 3, 3, 2,
     "Pipeline grew 2.6x in 11 days. Mixes chat handlers, ingest helpers, regenerate-rollback logic, attachment download. Will keep absorbing scope if not split.",
     "Split: chat handlers, ingest helpers, regenerate handler. Move OpenWebUI-specific regenerate logic to engine where FSM lives, not pipeline transport layer."),
    ("DBT-014", "MIME-type classification duplicated across 7 adapters", "mira-bots", "telegram/, slack/, teams/, gchat/, email/ chat_adapters", "Code", "Medium", "S", 3, 2, 1,
     "Each adapter re-implements mime->kind mapping. _IMAGE_MIMES copied 4+ places. Adding a new mime requires 7 edits.",
     "shared/chat/mime_utils.py with mime_to_kind(mime: str) -> Literal['image','pdf','other']. All adapters import."),
    ("DBT-015", "Image resize logic duplicated 3x with divergent MAX_VISION_PX", "mira-bots, mira-pipeline", "telegram/; whatsapp/bot.py:76; pipeline/main.py:75", "Code", "Medium", "S", 3, 2, 1,
     "_resize_for_vision implemented identically in 3+ places. MAX_VISION_PX = 1024 (pipeline) but 512 (whatsapp). Different image quality reaching vision models.",
     "shared/image_utils.py with resize_for_vision(img, max_px). Single MAX_VISION_PX constant. All callers import."),
    ("DBT-016", "Two adapter abstractions coexist (MIRAAdapter + ChatAdapter Protocol)", "mira-bots", "shared/adapters/base.py vs ChatAdapter usages", "Architecture", "Medium", "M", 3, 2, 3,
     "Email + WhatsApp extend MIRAAdapter base class. Telegram/Slack/Teams/GChat use ChatAdapter Protocol. Two parallel adapter patterns confuse maintainers.",
     "Migrate Email + WhatsApp to ChatAdapter Protocol. Deprecate MIRAAdapter. Easier after dormant adapters (DBT-022) are archived."),
    ("DBT-017", "Duplicate cmms_write_work_order + cmms_create_work_order tools", "mira-mcp", "server.py:270 + 300 (both @mcp.tool)", "Code", "High", "S", 4, 3, 1,
     "Both tools registered as MCP tools with identical signatures. cmms_create_work_order delegates to cmms_write_work_order with adapter fallback. Pollutes tool surface.",
     "Retire cmms_write_work_order. Keep cmms_create_work_order with adapter fallback. Update REST routes."),
    ("DBT-018", "Tool surface bloat: 10 CMMS tools with overlapping intent", "mira-mcp", "server.py lines 154-595 (10+ @mcp.tool decorators)", "Architecture", "Medium", "M", 3, 3, 2,
     "10 CMMS tools (write/create/list/complete WO; list/get assets; list PMs; health; create-from-nameplate). Clients must understand fallback semantics. Prompts bloat.",
     "Collapse into manage_cmms(action: str, params: dict). Returns union response. Reduces tool-search space ~5x."),
    ("DBT-019", "Schema drift between mcp SQLite and ingest NeonDB chunk shape", "mira-mcp, mira-core", "mcp/server.py:72-115 vs ingest/db/neon.py:345-404", "Architecture", "High", "L", 4, 4, 2,
     "mira-mcp stores equipment_photos with structured_description JSON. mira-core writes knowledge_entries with isa95_path/equipment_id/data_type. New field on one breaks recall on the other.",
     "Define KnowledgeChunk dataclass in shared module. Both ingest and recall validate. Document mapping photos -> knowledge_entries."),
    ("DBT-020", "NeonDB migrations as ad-hoc ensure_* functions in app code", "mira-core", "mira-core/mira-ingest/db/neon.py:137-184", "Architecture", "High", "M", 4, 4, 2,
     "ensure_image_embedding_column, ensure_knowledge_hierarchy_columns called from FastAPI startup. Catch-all Exception. Failures leave DB inconsistent. No audit of what ran.",
     "Move to db/migrations/ as numbered SQL files. Dedicated runner (Alembic or custom) at container startup. Log each migration outcome."),
    ("DBT-021", "mira-hub missing from CLAUDE.md repo map", "repo-wide", "CLAUDE.md (0 mentions of mira-hub)", "Architecture", "High", "S", 4, 3, 1,
     "14 in-flight feature branches per pre-merge-review. Module not in CLAUDE.md repo map. Onboarding agents have no signpost. Documentation drift.",
     "Add mira-hub entry to CLAUDE.md repo map: 'Next.js 15 CMMS rebuild (Atlas migration target). 14 feature branches in parallel. Not yet in main.'"),
    ("DBT-022", "Dormant adapter modules still building", "mira-bots", "mira-bots/{teams,whatsapp,reddit}/", "Architecture", "Medium", "S", 3, 2, 1,
     "Teams + WhatsApp dormant since March (per April 14 elegance audit). Reddit unclear if anyone uses. All still in compose, still build.",
     "Move to attic/. Comment out of docker-compose. Remove from imports. Restore from git history if Azure/Twilio onboarding ever happens."),
    ("DBT-023", "Elegance-audit cleanups not executed (paperclip, mira_copy, bravo, ignition)", "repo-wide", "paperclip/, mira_copy/, bravo/, ignition/", "Architecture", "Medium", "S", 3, 2, 1,
     "April 14 elegance audit recommended archiving these. mira-hud and mira-prototype were actioned but these four were not. Dead weight in the repo.",
     "git checkout -b archive/<dir>-2026-04 -- <dir>; git rm -r <dir>; commit. Document in CLAUDE.md Deferred / Archived table."),
    ("DBT-024", "Sidecar OEM migration: no owner, no ETA", "mira-sidecar", "tools/migrate_sidecar_oem_to_owui.py exists, never executed", "Architecture", "Medium", "L", 3, 3, 2,
     "ADR-0008 plans sidecar sunset. 398 OEM chunks must move to Open WebUI KB first. Runbook exists. No tracking issue. No owner. No ETA.",
     "File issue 'Archive mira-sidecar post-OEM migration'. Assign owner. Set 2026-06-01 target. Run runbook docs/runbooks/sidecar-oem-migration.md."),
    ("DBT-025", "known-issues.md stale on web cutover", "docs", "known-issues.md L17 vs mira-web/src/lib/mira-chat.ts:11", "Doc-Architecture", "Low", "S", 2, 2, 1,
     "known-issues.md says 'mira-web -> mira-pipeline cutover pending'. Verified: mira-chat.ts already calls pipeline:9099. Stale doc misleads agents.",
     "Delete the bullet from known-issues.md. Add CHANGELOG.md entry: 'mira-web now calls mira-pipeline :9099 (ADR-0008).'"),
    ("DBT-026", "39 containers across compose overlays; 9 in mira-core", "repo-wide", "docker-compose*.yml", "Architecture", "Medium", "XL", 3, 3, 3,
     "Per April 14 audit: 39 containers (4 compose topologies). 9 just in mira-core. Cognitive load + 39 healthchecks + 39 failure modes.",
     "Two compose profiles: 'core' (target 12-15 containers) and 'full' (everything). Long-term elegance bet. XL scoped to multiple sprints."),
    ("DBT-027", "Atlas (mira-cmms) vs mira-hub strategy unclear", "mira-cmms, mira-hub", "both have asset/work-order schemas", "Architecture", "High", "M", 4, 4, 2,
     "mira-cmms is Atlas (third-party Java/Postgres). mira-hub is parallel Next.js CMMS rebuild with own assets/work-orders/PMs. No ADR defines whether mira-hub replaces, integrates, or coexists.",
     "Write ADR-0032: 'mira-hub CMMS strategy.' Decide replacement vs integration. Document data-migration path and Atlas sunsetting timeline if applicable."),
    ("DBT-028", "mira-crawler 11.9K LOC mixes KB ingest + GTM marketing", "mira-crawler", "mira-crawler/tasks/* (11+ task files)", "Architecture", "Medium", "XL", 3, 3, 3,
     "Per April 14 audit: KB ingest tasks (discover, sitemaps, ingest, rss, gdrive) tangled with content-marketing tasks (blog, linkedin, social, content, youtube). Different products in same module.",
     "Split into mira-crawler (KB ingest) and mira-content (GTM marketing). Separate Celery queues. Long-term."),
]

# ---- Build workbook ----
wb = Workbook()

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="2F5496")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

SEV_FILLS = {
    "Critical": PatternFill("solid", start_color="F8CBAD"),
    "High":     PatternFill("solid", start_color="FFE699"),
    "Medium":   PatternFill("solid", start_color="FFF2CC"),
    "Low":      PatternFill("solid", start_color="E2EFDA"),
}
PHASE_FILLS = {
    1: PatternFill("solid", start_color="DDEBF7"),
    2: PatternFill("solid", start_color="EDEDED"),
    3: PatternFill("solid", start_color="F4F0EC"),
}

# ---- Sheet 1: Register ----
ws = wb.active
ws.title = "Register"
headers = ["ID", "Title", "Module", "File anchor", "Category", "Severity",
           "Effort", "Impact", "Risk", "Score", "Phase", "Why it matters", "Concrete fix"]
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

for row_idx, f in enumerate(FINDINGS, start=2):
    (id_, title, module, file_anchor, cat, sev, eff, imp, risk, phase, why, fix) = f
    ws.cell(row=row_idx, column=1, value=id_)
    ws.cell(row=row_idx, column=2, value=title)
    ws.cell(row=row_idx, column=3, value=module)
    ws.cell(row=row_idx, column=4, value=file_anchor)
    ws.cell(row=row_idx, column=5, value=cat)
    ws.cell(row=row_idx, column=6, value=sev)
    ws.cell(row=row_idx, column=7, value=eff)
    ws.cell(row=row_idx, column=8, value=imp)
    ws.cell(row=row_idx, column=9, value=risk)
    eff_lookup = f'IF(G{row_idx}="S",1,IF(G{row_idx}="M",2,IF(G{row_idx}="L",3,5)))'
    ws.cell(row=row_idx, column=10, value=f"=(H{row_idx}+I{row_idx})*(6-{eff_lookup})")
    ws.cell(row=row_idx, column=11, value=phase)
    ws.cell(row=row_idx, column=12, value=why)
    ws.cell(row=row_idx, column=13, value=fix)

    for col in range(1, 14):
        c = ws.cell(row=row_idx, column=col)
        c.font = body_font
        c.border = border
        c.alignment = center_align if col in (5, 6, 7, 8, 9, 10, 11) else wrap_align
    ws.cell(row=row_idx, column=6).fill = SEV_FILLS[sev]
    ws.cell(row=row_idx, column=11).fill = PHASE_FILLS[phase]

widths = [11, 38, 22, 38, 14, 11, 8, 8, 6, 8, 7, 60, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 28
for r in range(2, len(FINDINGS) + 2):
    ws.row_dimensions[r].height = 75

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{len(FINDINGS) + 1}"

ws.conditional_formatting.add(
    f"J2:J{len(FINDINGS) + 1}",
    CellIsRule(operator="greaterThanOrEqual", formula=["35"], fill=PatternFill("solid", start_color="C6EFCE"))
)

# ---- Per-phase sheets ----
def write_phase_sheet(workbook, phase_num, sheet_title, intro):
    sh = workbook.create_sheet(sheet_title)
    sh.append([intro])
    sh["A1"].font = Font(name="Arial", bold=True, size=12)
    sh.merge_cells("A1:F1")
    sh.append([])
    sub_headers = ["ID", "Title", "Severity", "Effort", "Score", "One-line fix"]
    sh.append(sub_headers)
    for col_idx in range(1, 7):
        c = sh.cell(row=3, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    items = [f for f in FINDINGS if f[9] == phase_num]
    items.sort(key=lambda x: (x[7] + x[8]) * (6 - EFFORT_NUM[x[6]]), reverse=True)

    for i, f in enumerate(items):
        (id_, title, _, _, _, sev, eff, imp, risk, _, _, fix) = f
        r = 4 + i
        sh.cell(row=r, column=1, value=id_)
        sh.cell(row=r, column=2, value=title)
        sh.cell(row=r, column=3, value=sev)
        sh.cell(row=r, column=4, value=eff)
        sh.cell(row=r, column=5, value=(imp + risk) * (6 - EFFORT_NUM[eff]))
        sh.cell(row=r, column=6, value=fix)
        for col in range(1, 7):
            c = sh.cell(row=r, column=col)
            c.font = body_font
            c.border = border
            c.alignment = wrap_align if col in (2, 6) else center_align
        sh.cell(row=r, column=3).fill = SEV_FILLS[sev]

    for i, w in enumerate([11, 42, 11, 8, 8, 60], start=1):
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.row_dimensions[1].height = 30
    sh.row_dimensions[3].height = 24
    for r in range(4, 4 + len(items)):
        sh.row_dimensions[r].height = 50
    sh.freeze_panes = "A4"
    sh.auto_filter.ref = f"A3:F{3 + len(items)}"

write_phase_sheet(wb, 1, "Phase 1 - Sprint",
                  "Phase 1 - Sprint (next 2 weeks): blast-radius reducers. 11/12 are S-effort. None block parity-sprint merges.")
write_phase_sheet(wb, 2, "Phase 2 - 90-day MVP",
                  "Phase 2 - 90-day MVP (through 2026-07-19): structural moves the parity-sprint merges and SOC 2 prep depend on.")
write_phase_sheet(wb, 3, "Phase 3 - Post-MVP",
                  "Phase 3 - Post-MVP (Q3+ 2026): big-ticket refactors. L/XL effort. Schedule before files cross next size threshold.")

# ---- Summary sheet ----
sh = wb.create_sheet("Summary")
sh["A1"] = "MIRA Tech Debt Audit - 2026-04-25 - Summary"
sh["A1"].font = Font(name="Arial", bold=True, size=14)
sh.merge_cells("A1:E1")
n_code = sum(1 for f in FINDINGS if f[4] == "Code")
n_arch = sum(1 for f in FINDINGS if f[4] == "Architecture")
n_doc = sum(1 for f in FINDINGS if f[4] == "Doc-Architecture")
sh["A2"] = f"Total findings: {len(FINDINGS)} - Code: {n_code}, Architecture: {n_arch}, Doc-Architecture: {n_doc}"
sh["A2"].font = Font(name="Arial", italic=True, size=10)
sh.merge_cells("A2:E2")

sh["A4"] = "By severity"
sh["A4"].font = Font(name="Arial", bold=True, size=11)
sh["A5"] = "Severity"; sh["B5"] = "Count"
for k in ("A5", "B5"):
    sh[k].font = header_font
    sh[k].fill = header_fill
    sh[k].border = border
sev_counts = [(s, sum(1 for f in FINDINGS if f[5] == s)) for s in ("Critical", "High", "Medium", "Low")]
for i, (sev, count) in enumerate(sev_counts, start=6):
    sh.cell(row=i, column=1, value=sev).fill = SEV_FILLS[sev]
    sh.cell(row=i, column=2, value=count)
    for col in (1, 2):
        c = sh.cell(row=i, column=col)
        c.font = body_font
        c.border = border
        c.alignment = center_align

sh["D4"] = "By phase"
sh["D4"].font = Font(name="Arial", bold=True, size=11)
sh["D5"] = "Phase"; sh["E5"] = "Count"
for k in ("D5", "E5"):
    sh[k].font = header_font
    sh[k].fill = header_fill
    sh[k].border = border
phase_labels = {1: "Phase 1 - Sprint", 2: "Phase 2 - 90-day MVP", 3: "Phase 3 - Post-MVP"}
for i, p in enumerate((1, 2, 3), start=6):
    sh.cell(row=i, column=4, value=phase_labels[p]).fill = PHASE_FILLS[p]
    sh.cell(row=i, column=5, value=sum(1 for f in FINDINGS if f[9] == p))
    for col in (4, 5):
        c = sh.cell(row=i, column=col)
        c.font = body_font
        c.border = border
        c.alignment = center_align

sh["A12"] = "By effort size"
sh["A12"].font = Font(name="Arial", bold=True, size=11)
sh["A13"] = "Effort"; sh["B13"] = "Count"; sh["C13"] = "Meaning"
for k in ("A13", "B13", "C13"):
    sh[k].font = header_font
    sh[k].fill = header_fill
    sh[k].border = border
eff_meaning = {"S": "<= 1 day", "M": "1-3 days", "L": "1-2 weeks", "XL": "sprint+"}
for i, eff in enumerate(("S", "M", "L", "XL"), start=14):
    sh.cell(row=i, column=1, value=eff)
    sh.cell(row=i, column=2, value=sum(1 for f in FINDINGS if f[6] == eff))
    sh.cell(row=i, column=3, value=eff_meaning[eff])
    for col in (1, 2, 3):
        c = sh.cell(row=i, column=col)
        c.font = body_font
        c.border = border
        c.alignment = center_align

sh["A20"] = "Top 5 by score (foundation-first sequencing applied)"
sh["A20"].font = Font(name="Arial", bold=True, size=11)
sh["A21"] = "ID"; sh["B21"] = "Title"; sh["C21"] = "Severity"; sh["D21"] = "Score"; sh["E21"] = "Phase"
for k in ("A21", "B21", "C21", "D21", "E21"):
    sh[k].font = header_font
    sh[k].fill = header_fill
    sh[k].border = border

def score(f):
    return (f[7] + f[8]) * (6 - EFFORT_NUM[f[6]])

top5 = sorted(FINDINGS, key=score, reverse=True)[:5]
for i, f in enumerate(top5, start=22):
    sh.cell(row=i, column=1, value=f[0])
    sh.cell(row=i, column=2, value=f[1])
    sh.cell(row=i, column=3, value=f[5]).fill = SEV_FILLS[f[5]]
    sh.cell(row=i, column=4, value=score(f))
    sh.cell(row=i, column=5, value=phase_labels[f[9]]).fill = PHASE_FILLS[f[9]]
    for col in (1, 2, 3, 4, 5):
        c = sh.cell(row=i, column=col)
        c.font = body_font
        c.border = border
        c.alignment = wrap_align if col == 2 else center_align

for col, w in zip("ABCDE", (16, 50, 14, 10, 24)):
    sh.column_dimensions[col].width = w
sh.row_dimensions[1].height = 22
for r in range(22, 27):
    sh.row_dimensions[r].height = 30

# Reorder so Summary is the first tab
wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"Findings: {len(FINDINGS)}")
